#!/usr/bin/env python3
"""Validação T-2 — atribuição automática de tabulação."""

from __future__ import annotations

import asyncio
import sys
import uuid
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import selectinload

_BACKEND = Path(__file__).resolve().parents[1]
_ROOT = _BACKEND.parent
for p in (_ROOT, _BACKEND, _ROOT / "worker"):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from app.core.database import AsyncSessionLocal
from app.models.campaign import Campaign
from app.models.lead import Lead
from app.models.lead_base import LeadBase
from app.models.lead_interaction import LeadInteraction
from app.models.tabulacao import Tabulacao
from app.models.user import User
from app.services.devolutiva import gerar_devolutiva_base
from app.services.tabulacao_assignment import apply_tabulacao, maybe_apply_tabulacao_on_transition
from worker.tasks.lead_tracking import track_inbound_lead_interaction


def _ok(label: str, cond: bool, detail: str = "") -> bool:
    status = "OK" if cond else "FALHA"
    print(f"  [{status}] {label}" + (f" — {detail}" if detail else ""))
    return cond


async def _ensure_neg_recusado(session) -> None:
    result = await session.execute(
        select(Tabulacao).where(Tabulacao.codigo == "NEG:RECUSADO")
    )
    if result.scalar_one_or_none() is None:
        from app.core.seed import seed_default_tabulacoes

        await seed_default_tabulacoes(session)


async def _fixture_lead_interaction(session) -> tuple[LeadInteraction, Lead, User]:
    user = (
        await session.execute(select(User).where(User.email == "admin@admin.com"))
    ).scalar_one()
    campaign = (
        await session.execute(select(Campaign).where(Campaign.user_id == user.id).limit(1))
    ).scalar_one_or_none()
    if campaign is None:
        raise RuntimeError("Nenhuma campanha encontrada para testes")

    lead_base = (
        await session.execute(
            select(LeadBase).where(LeadBase.campaign_id == campaign.id).limit(1)
        )
    ).scalar_one_or_none()
    if lead_base is None:
        raise RuntimeError("Nenhuma lead_base encontrada para testes")

    lead = (
        await session.execute(
            select(Lead).where(Lead.lead_base_id == lead_base.id).limit(1)
        )
    ).scalar_one_or_none()
    if lead is None:
        lead = Lead(
            user_id=user.id,
            lead_base_id=lead_base.id,
            id_cliente=f"T2-{uuid.uuid4().hex[:8]}",
            nome_cliente="Lead Tab T2",
            telefone_1="5511999999001",
        )
        session.add(lead)
        await session.flush()

    channel = "whatsapp"
    existing = (
        await session.execute(
            select(LeadInteraction).where(
                LeadInteraction.lead_id == lead.id,
                LeadInteraction.campaign_id == campaign.id,
                LeadInteraction.channel_type == channel,
            )
        )
    ).scalar_one_or_none()

    if existing:
        record = existing
        record.status = "em_andamento"
        record.tabulacao_id = None
        record.tabulacao_origem = None
        record.tabulacao_aplicada_em = None
        record.devolutiva = None
    else:
        record = LeadInteraction(
            lead_id=lead.id,
            campaign_id=campaign.id,
            channel_type=channel,
            status="em_andamento",
            tentativas=1,
            data_acionamento=datetime.now(timezone.utc),
        )
        session.add(record)

    await session.flush()
    await session.refresh(record, ["campaign"])
    return record, lead, user


async def test_rule_purchase(session) -> bool:
    print("\n=== Regra intent purchase → NEG:VENDA ===")
    record, lead, _ = await _fixture_lead_interaction(session)
    record.status = "em_andamento"
    record.tabulacao_id = None

    rec = await track_inbound_lead_interaction(
        session,
        "whatsapp",
        lead.telefone_1 or "5511999999001",
        "Quero comprar o produto agora",
        "purchase",
    )
    await session.commit()
    await session.refresh(rec, ["tabulacao"])

    codigo = rec.tabulacao.codigo if rec and rec.tabulacao else None
    return _ok(
        "purchase → NEG:VENDA",
        rec is not None
        and rec.tabulacao_origem == "INTENT"
        and codigo == "NEG:VENDA",
        f"origem={rec.tabulacao_origem if rec else None} codigo={codigo}",
    )


async def test_rule_nao_atendido(session) -> bool:
    print("\n=== Regra status nao_atendido → NEG:AUSENTE ===")
    record, _, _ = await _fixture_lead_interaction(session)
    record.status = "acionado"
    record.tabulacao_id = None
    record.tabulacao_origem = None

    applied = await maybe_apply_tabulacao_on_transition(
        session,
        record,
        status_interno="nao_atendido",
        channel=record.channel_type,
    )
    await session.commit()
    await session.refresh(record, ["tabulacao"])

    codigo = record.tabulacao.codigo if record.tabulacao else None
    return _ok(
        "nao_atendido → NEG:AUSENTE",
        applied and codigo == "NEG:AUSENTE" and record.tabulacao_origem == "INTENT",
        f"applied={applied} codigo={codigo} origem={record.tabulacao_origem}",
    )


async def test_no_tabulation_em_andamento(session) -> bool:
    print("\n=== Sem tabulação em em_andamento ===")
    record, lead, _ = await _fixture_lead_interaction(session)
    record.tabulacao_id = None

    rec = await track_inbound_lead_interaction(
        session,
        "whatsapp",
        lead.telefone_1 or "5511999999001",
        "Olá, só uma dúvida",
        "question",
    )
    await session.commit()
    return _ok(
        "em_andamento sem sinal → tabulacao_id None",
        rec is not None and rec.tabulacao_id is None,
        f"status={rec.status if rec else None} tab_id={rec.tabulacao_id if rec else None}",
    )


async def test_ia_classification(session) -> bool:
    print("\n=== IA: terminal sem regra (erro + número errado) ===")
    record, _, _ = await _fixture_lead_interaction(session)
    record.status = "erro"
    record.tabulacao_id = None
    record.tabulacao_origem = None
    record.devolutiva = (
        "Cliente informou que o número está errado e não conhece a empresa. "
        "Pediu para não ligar mais neste telefone."
    )

    applied = await apply_tabulacao(
        session,
        record,
        status_interno="erro",
        channel=record.channel_type,
        conversation_text=record.devolutiva,
    )
    await session.commit()
    await session.refresh(record, ["tabulacao"])

    if not applied or record.tabulacao is None:
        return _ok("IA classificou", False, "LLM não retornou tabulação válida")

    catalog = (
        await session.execute(select(Tabulacao.codigo))
    ).scalars().all()
    in_catalog = record.tabulacao.codigo in catalog
    return _ok(
        "IA com código no catálogo",
        record.tabulacao_origem == "IA" and in_catalog,
        f"codigo={record.tabulacao.codigo} origem={record.tabulacao_origem}",
    )


async def test_sip_hook(session) -> bool:
    print("\n=== Gancho SIP: SIP:486 → Ocupado ===")
    record, _, _ = await _fixture_lead_interaction(session)
    record.tabulacao_id = None

    applied = await apply_tabulacao(
        session,
        record,
        sip_code="SIP:486",
        status_interno=record.status,
    )
    await session.commit()
    await session.refresh(record, ["tabulacao"])

    nome = record.tabulacao.nome if record.tabulacao else None
    return _ok(
        "SIP:486 → Ocupado origem SIP",
        applied
        and record.tabulacao_origem == "SIP"
        and record.tabulacao
        and record.tabulacao.codigo == "SIP:486"
        and nome == "Ocupado",
        f"nome={nome} origem={record.tabulacao_origem}",
    )


async def test_devolutiva_column(session) -> bool:
    print("\n=== Devolutiva Excel — coluna Tabulação ===")
    record, lead, _ = await _fixture_lead_interaction(session)
    record.status = "convertido"
    record.devolutiva = "Venda confirmada"
    await apply_tabulacao(
        session,
        record,
        intent="purchase",
        status_interno="convertido",
        conversation_text=record.devolutiva,
    )
    await session.commit()

    lead_base_id = lead.lead_base_id
    xlsx_bytes = await gerar_devolutiva_base(session, lead_base_id)
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    tab_col = headers.index("Tabulação") + 1 if "Tabulação" in headers else None
    status_col = headers.index("Status operacional") + 1 if "Status operacional" in headers else None

    if tab_col is None:
        return _ok("coluna Tabulação presente", False, f"headers={headers}")

    lead_row = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == lead.id_cliente:
            lead_row = row
            break
    if lead_row is None:
        for row in range(2, sheet.max_row + 1):
            lead_row = row
            break

    tab_value = sheet.cell(row=lead_row, column=tab_col).value
    status_value = sheet.cell(row=lead_row, column=status_col).value if status_col else ""

    # Lead sem tabulação em outra linha — verificar célula vazia possível
    empty_ok = tab_value in ("", None) or tab_value == "Venda"

    return _ok(
        "Tabulação na planilha",
        tab_value == "Venda" and status_value == "Convertido",
        f"Tabulação={tab_value!r} Status operacional={status_value!r}",
    ) and _ok("headers incluem Categoria Tabulação", "Categoria Tabulação" in headers)


async def main() -> int:
    print("Validação T-2 — Tabulação automática")
    results: list[bool] = []

    async with AsyncSessionLocal() as session:
        await _ensure_neg_recusado(session)
        await session.commit()

    async with AsyncSessionLocal() as session:
        results.append(await test_rule_purchase(session))
        results.append(await test_rule_nao_atendido(session))
        results.append(await test_no_tabulation_em_andamento(session))
        results.append(await test_sip_hook(session))
        results.append(await test_devolutiva_column(session))

    async with AsyncSessionLocal() as session:
        results.append(await test_ia_classification(session))

    passed = sum(results)
    total = len(results)
    print(f"\n=== Resumo: {passed}/{total} cenários OK ===")
    return 0 if passed == total else 1


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
