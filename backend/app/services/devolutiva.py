"""Geração de planilhas de devolutiva (Excel) por base de leads."""

from __future__ import annotations

import re
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.lead import Lead
from app.models.lead_base import LeadBase
from app.models.lead_interaction import LeadInteraction

FIXED_COLUMNS: list[tuple[str, str]] = [
    ("id_cliente", "ID Cliente"),
    ("nome_cliente", "Nome"),
    ("cpf_cliente", "CPF"),
    ("email_cliente", "Email"),
    ("telefone_1", "Telefone 1"),
    ("telefone_2", "Telefone 2"),
    ("telefone_3", "Telefone 3"),
]

DEVOLUTIVA_COLUMNS: list[str] = [
    "Data Acionamento",
    "Status operacional",
    "Tabulação",
    "Categoria Tabulação",
    "Canal Atendimento",
    "Data Último Contato",
    "Devolutiva",
]

STATUS_LABELS: dict[str, str] = {
    "pendente": "Pendente",
    "acionado": "Acionado",
    "em_andamento": "Em andamento",
    "nao_atendido": "Não atendido",
    "convertido": "Convertido",
    "recusou": "Recusou",
    "erro": "Erro no acionamento",
}

CHANNEL_LABELS: dict[str, str] = {
    "whatsapp": "WhatsApp",
    "telegram": "Telegram",
    "voice": "Voz",
}

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True)

DEVOLUTIVAS_ROOT = Path("/workspace/devolutivas")
DATE_FILENAME_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _sort_aux_keys(column_mapping: dict[str, Any]) -> list[str]:
    return sorted(column_mapping.keys(), key=lambda key: int(key.replace("aux", "")))


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.astimezone().strftime("%d/%m/%Y %H:%M")


def _status_label(status: str | None) -> str:
    if not status:
        return ""
    return STATUS_LABELS.get(status, status)


def _channel_label(channel_type: str | None) -> str:
    if not channel_type:
        return ""
    return CHANNEL_LABELS.get(channel_type.lower(), channel_type)


async def _load_latest_interactions_by_lead(
    session: AsyncSession,
    lead_ids: list[uuid.UUID],
) -> dict[uuid.UUID, LeadInteraction]:
    if not lead_ids:
        return {}

    result = await session.execute(
        select(LeadInteraction)
        .options(selectinload(LeadInteraction.tabulacao))
        .where(LeadInteraction.lead_id.in_(lead_ids))
        .order_by(
            LeadInteraction.lead_id,
            LeadInteraction.data_ultimo_contato.desc().nullslast(),
        )
    )

    latest: dict[uuid.UUID, LeadInteraction] = {}
    for interaction in result.scalars().all():
        if interaction.lead_id not in latest:
            latest[interaction.lead_id] = interaction
    return latest


def _build_workbook(
    lead_base: LeadBase,
    leads: list[Lead],
    latest_by_lead: dict[uuid.UUID, LeadInteraction],
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Devolutiva"

    aux_keys = _sort_aux_keys(lead_base.column_mapping or {})
    headers = [label for _, label in FIXED_COLUMNS]
    headers.extend(lead_base.column_mapping.get(key, key) for key in aux_keys)
    headers.extend(DEVOLUTIVA_COLUMNS)

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, lead in enumerate(leads, start=2):
        col_index = 1
        for field_key, _ in FIXED_COLUMNS:
            sheet.cell(row=row_index, column=col_index, value=getattr(lead, field_key) or "")
            col_index += 1

        for aux_key in aux_keys:
            sheet.cell(row=row_index, column=col_index, value=lead.aux_values.get(aux_key, ""))
            col_index += 1

        interaction = latest_by_lead.get(lead.id)
        tab_nome = ""
        tab_categoria = ""
        if interaction and interaction.tabulacao:
            tab_nome = interaction.tabulacao.nome
            tab_categoria = interaction.tabulacao.categoria or ""
        elif interaction and interaction.tabulacao_id:
            tab_nome = "—"

        devolutiva_values = [
            _format_datetime(interaction.data_acionamento) if interaction else "",
            _status_label(interaction.status if interaction else "pendente"),
            tab_nome,
            tab_categoria,
            _channel_label(interaction.channel_type if interaction else ""),
            _format_datetime(interaction.data_ultimo_contato) if interaction else "",
            interaction.devolutiva if interaction and interaction.devolutiva else "",
        ]
        for value in devolutiva_values:
            sheet.cell(row=row_index, column=col_index, value=value)
            col_index += 1

    sheet.freeze_panes = "A2"

    for col_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_index)
        max_length = 0
        for cell in sheet[column_letter]:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)

    return workbook


async def gerar_devolutiva_base(session: AsyncSession, lead_base_id: uuid.UUID) -> bytes:
    """Gera xlsx de devolutiva para uma base de leads."""
    lead_base = await session.get(LeadBase, lead_base_id)
    if lead_base is None:
        raise ValueError(f"Lead base {lead_base_id} not found")

    leads_result = await session.execute(
        select(Lead)
        .where(Lead.lead_base_id == lead_base_id)
        .order_by(Lead.nome_cliente)
    )
    leads = list(leads_result.scalars().all())
    latest_by_lead = await _load_latest_interactions_by_lead(session, [lead.id for lead in leads])

    workbook = _build_workbook(lead_base, leads, latest_by_lead)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def list_historical_devolutivas(lead_base_id: uuid.UUID) -> list[dict[str, int | str]]:
    """Lista arquivos xlsx históricos gerados pelo job diário."""
    directory = DEVOLUTIVAS_ROOT / str(lead_base_id)
    if not directory.is_dir():
        return []

    files: list[dict[str, int | str]] = []
    for path in directory.glob("*.xlsx"):
        if not DATE_FILENAME_PATTERN.match(path.stem):
            continue
        files.append(
            {
                "data": path.stem,
                "filename": path.name,
                "size_bytes": path.stat().st_size,
            }
        )

    files.sort(key=lambda item: str(item["data"]), reverse=True)
    return files


def read_historical_devolutiva(lead_base_id: uuid.UUID, data: str) -> bytes:
    """Lê um arquivo xlsx histórico. Levanta FileNotFoundError se não existir."""
    if not DATE_FILENAME_PATTERN.match(data):
        raise ValueError(f"Invalid date format: {data}")

    path = DEVOLUTIVAS_ROOT / str(lead_base_id) / f"{data}.xlsx"
    if not path.is_file():
        raise FileNotFoundError(str(path))
    return path.read_bytes()
